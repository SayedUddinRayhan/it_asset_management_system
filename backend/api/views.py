from rest_framework.viewsets import ModelViewSet
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.filters import SearchFilter, OrderingFilter
from .models import Vendor, Department, Status, Category, Product, ProductDocument, TransferLog, RepairStatus, RepairLog
from .serializers import VendorSerializer, DepartmentSerializer, StatusSerializer, CategorySerializer, ProductDocumentSerializer, ProductSerializer, TransferLogSerializer, RepairStatusSerializer, RepairLogSerializer
from rest_framework.response import Response
from rest_framework import status

import io
import pandas as pd
from django.http import HttpResponse
from rest_framework.views import APIView
from reportlab.lib.pagesizes import landscape, A4
from reportlab.lib.units import cm
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter


class VendorViewSet(ModelViewSet):
    queryset = Vendor.objects.all().order_by("-created_at")
    serializer_class = VendorSerializer

    filter_backends = [SearchFilter, OrderingFilter]
    search_fields = ["unique_code", "name"]
    ordering_fields = ["name", "created_at"]

    def perform_destroy(self, instance):
        instance.is_active = False
        instance.save(update_fields=["is_active"])


class DepartmentViewSet(ModelViewSet):
    queryset = Department.objects.all().order_by("-created_at")
    serializer_class = DepartmentSerializer

    filter_backends = [SearchFilter, OrderingFilter]
    search_fields = ["unique_code", "name"]
    ordering_fields = ["name", "created_at"]

    def perform_destroy(self, instance):
        instance.is_active = False
        instance.save(update_fields=["is_active"])


class StatusViewSet(ModelViewSet):
    queryset = Status.objects.filter(is_active=True).order_by("name")
    serializer_class = StatusSerializer

    filter_backends = [SearchFilter, OrderingFilter]
    search_fields = ["unique_code", "name"]
    ordering_fields = ["name", "created_at"]

    def perform_destroy(self, instance):
        instance.is_active = False
        instance.save(update_fields=["is_active"])


class CategoryViewSet(ModelViewSet):
    queryset = Category.objects.filter(is_active=True).order_by("name")
    serializer_class = CategorySerializer

    filter_backends = [SearchFilter, OrderingFilter]
    search_fields = ["unique_code", "name"]
    ordering_fields = ["name", "created_at"]

    def perform_destroy(self, instance):
        instance.is_active = False
        instance.save(update_fields=["is_active"])

class ProductViewSet(ModelViewSet):
    queryset = Product.objects.filter(is_active=True).select_related(
        "vendor", "current_department", "status"
    ).prefetch_related("documents").order_by("-created_at")
    serializer_class = ProductSerializer
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = ["status", "category", "current_department"]
    search_fields = ["unique_code", "name", "vendor__name", "current_department__name", "warranty_years"]
    ordering_fields = ["created_at", "name", "price"]

    def perform_destroy(self, instance):
        instance.is_active = False
        instance.save(update_fields=["is_active"])

    def create(self, request, *args, **kwargs):
        files = request.FILES.getlist("documents")

        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        in_use_status, _ = Status.objects.get_or_create(name="In Stock")
        product = serializer.save(status=in_use_status)


        for f in files:
            ProductDocument.objects.create(product=product, file=f)

        # re-serialize
        output_serializer = self.get_serializer(product)

        return Response(output_serializer.data, status=status.HTTP_201_CREATED)
    
    def partial_update(self, request, *args, **kwargs):
        product = self.get_object()
        status_id = request.data.get("status")

        if status_id:
            try:
                new_status = Status.objects.get(id=status_id)
                product.status = new_status
                product.save(update_fields=["status"])
            except Status.DoesNotExist:
                return Response({"error": "Invalid status"}, status=status.HTTP_400_BAD_REQUEST)

        return super().partial_update(request, *args, **kwargs)


class ProductDocumentViewSet(ModelViewSet):
    serializer_class = ProductDocumentSerializer

    def get_queryset(self):
        product_id = self.kwargs.get("product_id")
        return ProductDocument.objects.filter(product_id=product_id).order_by("-uploaded_at")


# Excel Export
class ProductExportExcelView(APIView):

    def post(self, request, *args, **kwargs):
        filters = request.data
        qs = Product.objects.filter(is_active=True)

        # Apply filters
        search = filters.get("search")
        status = filters.get("status")
        category = filters.get("category")
        department = filters.get("department")
        ordering = filters.get("ordering", "-created_at")

        if search:
            qs = qs.filter(unique_code__icontains=search) | qs.filter(name__icontains=search)
        if status:
            qs = qs.filter(status_id=status)
        if category:
            qs = qs.filter(category_id=category)
        if department:
            qs = qs.filter(current_department_id=department)

        qs = qs.order_by(ordering)

       
        data = []
        for p in qs:
            data.append({
                "ID": p.unique_code,
                "Name": p.name,
                "Model Number": p.model_number,
                "Serial Number": p.serial_number,
                "Description": p.description,
                "Category": p.category.name if p.category else "",
                "Department": p.current_department.name if p.current_department else "",
                "Vendor": p.vendor.name if p.vendor else "",
                "Price": float(p.price),
                "Quantity": p.quantity,
                "Purchase Date": p.purchase_date.strftime("%d-%m-%Y") if p.purchase_date else "",
                "Warranty": f"{p.warranty_years} years" if p.warranty_years else "",
                "Warranty End": p.warranty_end_date.strftime("%d-%m-%Y") if p.warranty_end_date else "",
                "Status": p.status.name if p.status else "",
                "Created At": p.created_at.strftime("%d-%m-%Y %H:%M:%S"),
                "Updated At": p.updated_at.strftime("%d-%m-%Y %H:%M:%S"),
            })

        df = pd.DataFrame(data)

        # Create Excel
        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Products")
            sheet = writer.book["Products"]

            # Style header
            header_font = Font(bold=True)
            for cell in sheet[1]:
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

            # Auto column width
            for col in sheet.columns:
                max_length = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    try:
                        max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                sheet.column_dimensions[col_letter].width = max_length + 4

            # Freeze header
            sheet.freeze_panes = "A2"

        buffer.seek(0)

        response = HttpResponse(
            buffer,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = "attachment; filename=products.xlsx"
        return response




# Export PDF
class ProductExportPDFView(APIView):
    def post(self, request, *args, **kwargs):
        filters = request.data
        qs = Product.objects.filter(is_active=True)

        # Apply filters
        search = filters.get("search")
        status = filters.get("status")
        category = filters.get("category")
        department = filters.get("department")
        ordering = filters.get("ordering", "-created_at")

        if search:
            qs = qs.filter(unique_code__icontains=search) | qs.filter(name__icontains=search)
        if status:
            qs = qs.filter(status_id=status)
        if category:
            qs = qs.filter(category_id=category)
        if department:
            qs = qs.filter(current_department_id=department)

        qs = qs.order_by(ordering)

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(A4),
            leftMargin=1*cm,
            rightMargin=1*cm,
            topMargin=2*cm,
            bottomMargin=1*cm
        )

        styles = getSampleStyleSheet()
        elements = []

        # Title
        title_style = styles["Title"]
        title_style.fontSize = 18
        title_style.leading = 22
        title = Paragraph("Products Report", title_style)
        elements.append(title)
        elements.append(Spacer(1, 12))

        # Table headers
        headers = ["SL", "Name", "Category", "Department", "Vendor", "Price", "Purchase", "Warranty", "End", "Status"]
        data = [headers]

        # Wrap long text using Paragraph
        wrap_style = styles["BodyText"]
        wrap_style.fontSize = 9
        wrap_style.leading = 11

        for i, prod in enumerate(qs, start=1):
            row = [
                str(i),
                Paragraph(prod.name if prod.name else "-", wrap_style),
                Paragraph(prod.category.name if prod.category else "-", wrap_style),
                Paragraph(prod.current_department.name if prod.current_department else "-", wrap_style),
                Paragraph(prod.vendor.name if prod.vendor else "-", wrap_style),
                f"{prod.price:.2f}",
                prod.purchase_date.strftime("%d-%m-%Y") if prod.purchase_date else "-",
                f"{prod.warranty_years}y" if prod.warranty_years else "-",
                prod.warranty_end_date.strftime("%d-%m-%Y") if prod.warranty_end_date else "-",
                Paragraph(prod.status.name if prod.status else "-", wrap_style),
            ]
            data.append(row)

        # Column widths (compact and balanced)
        col_widths = [
            1.0*cm,   # SL
            5.5*cm,   # Name
            3.0*cm,   # Category
            3.0*cm,   # Department
            3.0*cm,   # Vendor
            2.0*cm,   # Price
            2.5*cm,   # Purchase
            1.9*cm,   # Warranty
            2.0*cm,   # End
            2.3*cm,   # Status
        ]

        table = Table(data, colWidths=col_widths, repeatRows=1)

        # Professional table style
        style = TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f0f0f0")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 11),
            ("FONTSIZE", (0, 1), (-1, -1), 9),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
            ("LEFTPADDING", (0, 0), (-1, -1), 3),
            ("RIGHTPADDING", (0, 0), (-1, -1), 3),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ])

        # Alternating row colors
        for i in range(1, len(data)):
            if i % 2 == 0:
                style.add("BACKGROUND", (0, i), (-1, i), colors.HexColor("#fafafa"))

        table.setStyle(style)
        elements.append(table)

        # Header & Footer
        def header_footer(canvas_obj, doc_obj):
            canvas_obj.saveState()
            canvas_obj.setFont("Helvetica-Bold", 14)
            canvas_obj.drawString(2*cm, doc_obj.pagesize[1] - 1.5*cm, "Feni Diabetes Hospital")
            canvas_obj.setFont("Helvetica", 9)
            canvas_obj.drawRightString(doc_obj.pagesize[0] - 2*cm, 1*cm, f"Page {doc_obj.page}")
            canvas_obj.restoreState()

        # Build PDF
        doc.build(elements, onFirstPage=header_footer, onLaterPages=header_footer)

        buffer.seek(0)
        response = HttpResponse(buffer, content_type="application/pdf")
        response["Content-Disposition"] = "attachment; filename=products.pdf"
        return response



class TransferLogViewSet(ModelViewSet):
    queryset = TransferLog.objects.select_related(
        "product", "from_department", "to_department"
    ).order_by("-created_at")
    serializer_class = TransferLogSerializer

    def perform_create(self, serializer):
        transfer = serializer.save()
        if transfer.product:
            transfer.product.current_department = transfer.to_department
            transfer.product.save()

class RepairStatusViewSet(ModelViewSet):
    queryset = RepairStatus.objects.filter(is_active=True).order_by("name")
    serializer_class = RepairStatusSerializer

    filter_backends = [SearchFilter, OrderingFilter]
    search_fields = ["name"]
    ordering_fields = ["name", "created_at"]

    def perform_destroy(self, instance):
        instance.is_active = False
        instance.save(update_fields=["is_active"])


class RepairLogViewSet(ModelViewSet):
    queryset = RepairLog.objects.select_related(
        "product", "status"
    ).order_by("-created_at")
    serializer_class = RepairLogSerializer

    def perform_update(self, serializer):
        repair = serializer.save()

        if repair.product and repair.status and repair.status.product_status:
            repair.product.status = repair.status.product_status
            repair.product.save()
